import os
import typing
import openpyxl.styles as styles
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext

from bot import keyboards, models, filters
from datetime import datetime
from aiogram.types import FSInputFile
from aiogram.types import CallbackQuery
from aiogram import Dispatcher, F, types
from openpyxl import Workbook
from sqlalchemy import select


EXCEL_FILE = 'Matvei_bot'
LIST_SUBJECTS = ["index", "id", "first_name", "last_name", "username"]


async def run_db_export_handler(message: types.Message, state: FSMContext, session):
    await state.clear()

    date_time_2 = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    xlsx_file = "Matvei_bot" + date_time_2 + ".xlsx"

    async with session() as open_session:
        users_from_db = await open_session.execute(select(models.sql.User))
        users: typing.List[models.sql.User] = users_from_db.scalars().all()
    wb = Workbook()
    ws = wb.active

    for col_num, column_title in enumerate(LIST_SUBJECTS, 1):
        col_letter = ws.cell(row=1, column=col_num)
        col_letter.value = column_title

    for row_num, user in enumerate(users, 2):
        for col_num, column_title in enumerate(LIST_SUBJECTS, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = getattr(user, column_title.lower())
            cell.alignment = styles.Alignment(horizontal='center', vertical='center')
            cell.font = styles.Font(bold=True)
            cell.fill = styles.PatternFill(fill_type='solid', start_color='EEEEEE')

    wb.save(xlsx_file)

    await message.answer_document(document=FSInputFile(f"{xlsx_file}"),
                                           reply_markup=keyboards.inline.admin.admin_menu.keyboard.as_markup())
    os.remove(xlsx_file)


def setup(dp: Dispatcher):
    dp.message.register(
        run_db_export_handler,
        Command(commands="export"), filters.IsBotAdminFilter(is_bot_admin=True)
    )
